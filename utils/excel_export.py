"""
Excel Export Utility
Provides functions to export data to Excel with auto-formatted columns and rows
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def auto_adjust_column_width(worksheet, extra_width=2):
    """
    Auto-adjust column widths based on content
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + extra_width, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_excel_export(data, headers, title="Export", sheet_name="Data"):
    """
    Create a formatted Excel file from data
    
    Args:
        data: List of dictionaries or list of lists
        headers: List of header names
        title: Title for the export
        sheet_name: Name of the worksheet
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Timestamp row
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    timestamp_cell = ws['A2']
    timestamp_cell.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    timestamp_cell.font = Font(size=10, italic=True)
    timestamp_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    
    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[4].height = 25
    
    # Data rows
    for row_num, row_data in enumerate(data, 5):
        # Handle both dictionary and list formats
        if isinstance(row_data, dict):
            values = [row_data.get(header, '') for header in headers]
        else:
            values = row_data
        
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            
            # Alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws, extra_width=3)
    
    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A5'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_resident_list_export(residents):
    """Export resident list"""
    headers = ['Account Number', 'Name', 'Email', 'Unit', 'Enrollment Status', 'Monthly Rent', 'Lease Start']
    data = []
    
    for resident in residents:
        enrollment_status = resident.get('enrollment_status', 'Not Enrolled')
        # Convert to title case for display
        if enrollment_status == 'enrolled':
            enrollment_status = 'Enrolled'
        elif enrollment_status == 'not_enrolled':
            enrollment_status = 'Not Enrolled'
        
        data.append({
            'Account Number': resident.get('account_number', f"ACC2024{resident.get('id', ''):06d}"),
            'Name': resident.get('name', ''),
            'Email': resident.get('email', ''),
            'Unit': resident.get('unit', ''),
            'Enrollment Status': enrollment_status,
            'Monthly Rent': f"${resident.get('monthly_rent', 0):,.2f}",
            'Lease Start': resident.get('lease_start_date', 'N/A')
        })
    
    return create_excel_export(data, headers, 
                              title="Resident Rent Reporting List",
                              sheet_name="Residents")


def create_reporting_runs_export(runs):
    """Export reporting runs"""
    headers = ['Run ID', 'Date', 'Type', 'Status', 'Records', 'Success Rate', 'Notes']
    data = []
    
    for run in runs:
        data.append({
            'Run ID': run.get('id', ''),
            'Date': run.get('date', ''),
            'Type': run.get('type', ''),
            'Status': run.get('status', ''),
            'Records': run.get('records', ''),
            'Success Rate': run.get('success_rate', ''),
            'Notes': run.get('notes', '')
        })
    
    return create_excel_export(data, headers,
                              title="Metro2 Reporting Runs",
                              sheet_name="Reporting Runs")


def create_disputes_export(disputes):
    """Export disputes"""
    headers = ['ID', 'Date Submitted', 'Resident', 'Issue', 'Status', 'Priority', 'Details']
    data = []
    
    for dispute in disputes:
        data.append({
            'ID': dispute.get('id', ''),
            'Date Submitted': dispute.get('date', ''),
            'Resident': dispute.get('resident', ''),
            'Issue': dispute.get('issue', ''),
            'Status': dispute.get('status', ''),
            'Priority': dispute.get('priority', ''),
            'Details': dispute.get('details', '')
        })
    
    return create_excel_export(data, headers,
                              title="Rent Reporting Disputes",
                              sheet_name="Disputes")


def create_audit_logs_export(logs):
    """Export audit logs"""
    headers = ['ID', 'Timestamp', 'User', 'Action', 'Details']
    data = []
    
    for log in logs:
        data.append({
            'ID': log.get('id', ''),
            'Timestamp': log.get('timestamp', ''),
            'User': log.get('user', ''),
            'Action': log.get('action', ''),
            'Details': log.get('details', '')
        })
    
    return create_excel_export(data, headers,
                              title="System Audit Logs",
                              sheet_name="Audit Logs")
